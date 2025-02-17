# # Importando as bibliotecas necessárias
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
from tkinter import *

import os
import time
import datetime
import tkinter as tk

# Nome do arquivo Excel
arquivo_excel = "historico_temperatura.xlsx"

class ConsultaClima:

    # Configurando a tela da aplicação
    def __init__(self):        
        self.janela = tk.Tk()
        self.janela.title("Previsão do Tempo")
        self.janela.geometry("450x150")
        self.janela.resizable(False, False)

        self.label = tk.Label(self.janela, text="Digite a cidade para pesquisar o clima:", font=("Arial", 11))
        self.label.pack(pady=5)
        
        self.cidade_entry = tk.Entry(self.janela, font=("Arial", 10))
        self.cidade_entry.pack(pady=10)

        self.botao = tk.Button(
            self.janela,
            text="Buscar previsão",
            command=self.atualizar_previsao,
            font=("Arial", 10),
            width=20
        )
        self.botao.pack(pady=10)

    # Consulta para obter as informações do clima    
    def obter_clima(self, cidade):
        try: 
            # Configurações para acessar o navegador
            options = webdriver.ChromeOptions()
            options.add_argument("--disable-blink-features=AutomationControlled")  
            driver = webdriver.Chrome(options=options)

            # Acessando o Google
            driver.get("https://www.google.com")

            # Pesquisando o clima da cidade informada
            barra_pesquisa = driver.find_element(By.NAME, "q")
            barra_pesquisa.send_keys(f"Temperatura em {cidade}")
            time.sleep(1)
            barra_pesquisa.send_keys(Keys.RETURN)
            time.sleep(2)

            # Trecho onde obtém as informações de temperatura e Umidade
            temperatura = driver.find_element(By.ID, "wob_tm").text
            umidade_texto = driver.find_element(By.ID, "wob_hm").text
            umidade_valor = int(umidade_texto.strip('%'))

            if 12 <= umidade_valor <= 40:
                status_umidade = "Alerta Laranja, Baixa Umidade no ar"
            elif 41 <= umidade_valor <= 70:
                status_umidade = "Alerta Amarelo, Média Umidade no ar"
            elif umidade_valor >= 71:
                status_umidade = "Alta Umidade no ar"
            else:
                status_umidade = "Alerta Vermelho, Umidade extremamente baixa no ar"

            driver.quit()

            data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            return temperatura, umidade_texto, status_umidade, data_hora
        except:
            return None, None, "Erro ao obter os dados", None
    
    # Função para salvar as informações da consulta no Excel
    def salvar_em_excel(self, temperatura, umidade, status_umidade, data_hora, cidade):

        # Verifica se o arquivo existe
        if os.path.exists(arquivo_excel):
            # Carrega planilha existente
            wb = load_workbook(arquivo_excel)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Adicionando Cabeçalho na planilha nova
            ws.append(["Data / Hora", "Cidade", "Temperatura (\u00b0C)", "Umidade (%)", "Status Umidade"])

        # Adicionando novos dados na planilha
        ws.append([data_hora, cidade, temperatura, umidade, status_umidade])

        # Salvando o arquivo Excel
        wb.save(arquivo_excel)

    # Responsável por dar start na aplicação ao clicar no botão
    def atualizar_previsao(self):
        cidade = self.cidade_entry.get().strip()
        if not cidade:
            print("Por favor, digite uma cidade.")
            return
        
        # Executando a consulta de clima
        temperatura, umidade, status_umidade, data_hora = self.obter_clima(cidade)

        if temperatura and umidade:
            self.salvar_em_excel(temperatura, umidade, status_umidade, data_hora, cidade)
            print(f"Dados salvos para {cidade}: Temperatura {temperatura}\u00b0C, Umidade {umidade}, Status: {status_umidade}")
            mensagem = f"Dados salvos para {cidade}:\nTemperatura: {temperatura}°C\nUmidade: {umidade}\nStatus: {status_umidade}"
            messagebox.showinfo("Consulta Realizada", mensagem)
        else:
            print("Erro ao coletar os dados.")

    def iniciar(self):
        "Inicia o loop da interface"
        self.janela.mainloop()

# Instanciando e iniciando a tela da aplicação
if __name__ == "__main__":
    app = ConsultaClima()
    app.iniciar()